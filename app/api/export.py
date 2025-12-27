import re
import os
import logging
import zipfile
import io
import csv
import json
import hashlib
import requests # Added for image validation
import concurrent.futures
import time
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Dict, Any
from pydantic import BaseModel
from app.schemas.core import ModelPricingSnapshotResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.database import get_db
from app.models.core import Model, Series, Manufacturer, EquipmentType, ModelPricingSnapshot
from app.models.templates import AmazonProductType, ProductTypeField, EquipmentTypeProductType

from app.schemas.core import ModelPricingSnapshotResponse
from app.api.pricing import recalculate_targeted, PricingRecalculateRequest
from app.services.pricing_calculator import PricingConfigError

router = APIRouter(prefix="/export", tags=["export"])

# Module-level cache for HTTP results
# Structure: { url: { 'time': float, 'code': int|None, 'error': str|None } }
HTTP_CACHE = {}
MAX_CACHE_ENTRIES = 5000
TTL_SECONDS = 900 # 15 minutes

# --------------------------------------------------------------------------
# Staleness Check Helper
# --------------------------------------------------------------------------
def ensure_models_fresh_for_export(request: ExportPreviewRequest, db: Session) -> tuple[bool, int]:
    """
    Check if any selected model has stale pricing (or missing baseline snapshots).
    If so, trigger a targeted recalculation before export proceeds.
    Returns (recalc_performed: bool, recalc_model_count: int).
    """
    if not request.model_ids:
        return False, 0
        
    recalc_req = PricingRecalculateRequest(
        model_ids=request.model_ids,
        only_if_stale=True
    )
    
    # Direct service logic reuse via the API function, 
    # but we must ensure we don't double-wrap or mis-handle dependencies.
    # calling the route function directly is acceptable in FastAPI if deps match.
    # alternatively, extract the logic. 
    # For chunks, calling the route function is safest "minimal change" 
    # as long as we pass 'db'.
    
    try:
        resp = recalculate_targeted(recalc_req, db)
    except PricingConfigError as e:
        # If the underlying recalc fails due to config (e.g. fixed cell missing),
        # we must abort export and tell the user.
        raise HTTPException(status_code=400, detail=f"Export failed during pricing check: {str(e)}")
    
    return (resp.recalculated_models > 0), resp.recalculated_models

@router.get("/debug-price/{model_id}", response_model=ModelPricingSnapshotResponse)
def get_debug_price(model_id: int, db: Session = Depends(get_db)):
    """
    Get the baseline pricing snapshot for this model (Amazon / Choice No Padding).
    Used for UI debugging and validation.
    """
    snap = db.query(ModelPricingSnapshot).filter(
        ModelPricingSnapshot.model_id == model_id,
        ModelPricingSnapshot.marketplace == "amazon",
        ModelPricingSnapshot.variant_key == "choice_no_padding"
    ).first()
    
    if not snap:
        raise HTTPException(status_code=404, detail="Baseline snapshot not found. Please recalculate pricing or run seed.")
        
    return snap

class ExportPreviewRequest(BaseModel):
    model_ids: List[int]
    listing_type: str = "individual"  # "individual" or "parent_child"

class ExportRowData(BaseModel):
    model_id: int
    model_name: str
    data: List[str | None]

class ExportPreviewResponse(BaseModel):
    headers: List[List[str | None]]
    rows: List[ExportRowData]
    template_code: str
    export_signature: str
    field_map: Dict[str, int]
    audit: Dict[str, Any]
    recalc_performed: bool = False
    recalc_model_count: int = 0

def compute_export_signature(template_code: str, header_rows: list, data_rows: list) -> str:
    """
    Compute a deterministic SHA256 signature for the export content.
    Normalization Rules: None -> "", all values to strings, no trimming.
    """
    def normalize_row(row):
        return [str(val) if val is not None else "" for val in row]

    payload = {
        "template_code": template_code,
        "headers": [normalize_row(r) for r in header_rows],
        "rows": [normalize_row(r) for r in data_rows]
    }
    
    # Serialize to JSON with sorted keys to ensure determinism if we expanded payload
    # For now, keys are fixed. separators=(",", ":") removes whitespace.
    serialized = json.dumps(payload, separators=(",", ":"), ensure_ascii=False)
    
    return hashlib.sha256(serialized.encode('utf-8')).hexdigest()

def is_price_field(field_name: str) -> bool:
    """Exact logic from get_field_value triggering price population."""
    field_name_lower = field_name.lower()
    return "purchasable_offer[marketplace_id=atvpdkikx0der]" in field_name_lower and \
           "our_price#1.schedule#1.value_with_tax" in field_name_lower

def is_sku_related_field(field_name: str) -> bool:
    """True if field relates to SKU logic."""
    lower = field_name.lower()
    keys = ["item_sku", "contribution_sku", "parent_sku", "parent_child", "relationship_type"]
    return any(k in lower for k in keys)

def build_audit_columns(ordered_field_names: List[str], ordered_required_flags: List[bool]) -> List[dict]:
    """Identify which columns correspond to price and sku fields."""
    audited = []
    for idx, (name, required) in enumerate(zip(ordered_field_names, ordered_required_flags)):
        if not required:
            # Skip audit for fields that are not exported/populated
            continue
            
        if is_price_field(name):
            audited.append({"field_name": name, "column_index": idx, "type": "PRICE"})
        elif is_sku_related_field(name):
            audited.append({"field_name": name, "column_index": idx, "type": "SKU"})
    return audited

# Example usage:
# signature = compute_export_signature(code, headers, rows)

def get_field_source_audit(field: ProductTypeField, model: Model, series, manufacturer, equipment_type, is_image_field: bool) -> dict | None:
    """Determine the source of a field value for audit purposes."""
    if not field.required:
        return None
        
    lower = field.field_name.lower()
    # Focus audit on specific content fields prone to template defaults
    target_keys = ["product_description", "bullet_point", "generic_keyword"]
    if not any(k in lower for k in target_keys):
        return None

    if field.selected_value:
        val = substitute_placeholders(field.selected_value, model, series, manufacturer, equipment_type, is_image_url=is_image_field)
        return {"field_name": field.field_name, "source": "selected_value", "preview": val[:100]}
        
    if field.custom_value:
         val = substitute_placeholders(field.custom_value, model, series, manufacturer, equipment_type, is_image_url=is_image_field)
         return {"field_name": field.field_name, "source": "custom_value", "preview": val[:100]}
    
    return {"field_name": field.field_name, "source": "none", "preview": ""}

class ExportValidationIssue(BaseModel):
    severity: str  # "error", "warning"
    model_id: int | None = None
    model_name: str | None = None
    message: str

class ExportValidationResponse(BaseModel):
    status: str  # "valid", "warnings", "errors"
    summary_counts: Dict[str, int]
    items: List[ExportValidationIssue]

@router.post("/validate", response_model=ExportValidationResponse)
def validate_export(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """
    Pre-flight check for export readiness. Verifies templates, pricing snapshots, and data integrity.
    """
    issues = []
    
    # 1. Basic Request Validation
    if not request.model_ids:
        return ExportValidationResponse(
            status="errors",
            summary_counts={"total_models": 0, "issues": 1},
            items=[ExportValidationIssue(severity="error", message="No models selected.")]
        )

    models = db.query(Model).filter(Model.id.in_(request.model_ids)).all()
    if not models or len(models) != len(request.model_ids):
        found_ids = {m.id for m in models}
        missing_ids = set(request.model_ids) - found_ids
        issues.append(ExportValidationIssue(severity="error", message=f"Some requested models exist. Missing IDs: {missing_ids}"))

    total_models = len(models)
    
    image_fields = []
    
    # 2. Equipment Type Consistency & Template Loading
    if models:
        eq_ids = set(m.equipment_type_id for m in models)
        if len(eq_ids) > 1:
             issues.append(ExportValidationIssue(severity="error", message="Selected models have mixed Equipment Types. Export requires a single Equipment Type."))
        elif len(eq_ids) == 1:
             eq_id = list(eq_ids)[0]
             link = db.query(EquipmentTypeProductType).filter(EquipmentTypeProductType.equipment_type_id == eq_id).first()
             if not link:
                 equip = db.query(EquipmentType).filter(EquipmentType.id == eq_id).first()
                 name = equip.name if equip else f"ID {eq_id}"
                 issues.append(ExportValidationIssue(severity="error", message=f"No Amazon Template linked to Equipment Type: {name}"))
             else:
                 pt = db.query(AmazonProductType).filter(AmazonProductType.id == link.product_type_id).first()
                 if not pt or not pt.code:
                     issues.append(ExportValidationIssue(severity="error", message="Linked Amazon Template is invalid or missing Template Code."))
                 else:
                     # Load fields for placeholder verification
                     all_fields = db.query(ProductTypeField).filter(ProductTypeField.product_type_id == pt.id).all()
                     image_fields = [f for f in all_fields if is_image_url_field(f.field_name)]

    # 3. Model-Specific Checks
    MAX_HTTP_MODELS = 25
    MAX_HTTP_URLS = 60
    MAX_CONCURRENCY = 6

    http_models_checked_count = 0
    http_urls_scheduled_count = 0
    is_capped = False
    
    # Prune cache if needed
    if len(HTTP_CACHE) > MAX_CACHE_ENTRIES:
        HTTP_CACHE.clear()
        
    verification_jobs = [] # List of dicts: {'model': model, 'url': url, 'cached': bool}

    for idx, model in enumerate(models):
        # A. Pricing Snapshot
        snap = db.query(ModelPricingSnapshot).filter(
            ModelPricingSnapshot.model_id == model.id,
            ModelPricingSnapshot.marketplace == "amazon",
            ModelPricingSnapshot.variant_key == "choice_no_padding"
        ).first()
        
        if not snap:
            issues.append(ExportValidationIssue(
                severity="error", 
                model_id=model.id, 
                model_name=model.name, 
                message="Missing pricing snapshot (choice_no_padding). Recalculation required."
            ))
            
        # B. Image Placeholders & HTTP Check
        if image_fields:
            series = db.query(Series).filter(Series.id == model.series_id).first()
            manufacturer = db.query(Manufacturer).filter(Manufacturer.id == series.manufacturer_id).first() if series else None
            equip_type = db.query(EquipmentType).filter(EquipmentType.id == model.equipment_type_id).first()
            
            # 1. Placeholder Syntax Check (All Fields, Always)
            for img_field in image_fields:
                val_to_check = img_field.selected_value or img_field.custom_value
                
                if val_to_check:
                    resolved = substitute_placeholders(val_to_check, model, series, manufacturer, equip_type, is_image_url=True)
                    if '[' in resolved and ']' in resolved:
                         issues.append(ExportValidationIssue(
                            severity="warning",
                            model_id=model.id,
                            model_name=model.name,
                            message=f"Unresolved placeholder in image field '{img_field.field_name}': {resolved}"
                        ))
            
            # 2. HTTP Availability Check (Sampled, Capped, Cached)
            # Strategy: Full check for first model; First + Last fields only for subsequent models
            fields_to_check = image_fields if idx == 0 else ([image_fields[0], image_fields[-1]] if len(image_fields) > 1 else image_fields)
            
            model_triggering_fetch = False
            
            for img_field in fields_to_check:
                val = img_field.selected_value or img_field.custom_value
                if val:
                    url = substitute_placeholders(val, model, series, manufacturer, equip_type, is_image_url=True)
                    
                    if '[' in url and ']' in url:
                        continue
                    
                    # Check Cache
                    now = time.time()
                    cached_entry = HTTP_CACHE.get(url)
                    if cached_entry and (now - cached_entry['time'] < TTL_SECONDS):
                        verification_jobs.append({'model': model, 'url': url, 'cached': True, 'entry': cached_entry})
                        continue
                        
                    # Not Cached: Apply Caps
                    if is_capped:
                        continue
                        
                    if http_urls_scheduled_count >= MAX_HTTP_URLS:
                        is_capped = True
                        continue
                        
                    if not model_triggering_fetch and http_models_checked_count >= MAX_HTTP_MODELS:
                        is_capped = True
                        continue

                    # Schedule Fetch
                    verification_jobs.append({'model': model, 'url': url, 'cached': False})
                    http_urls_scheduled_count += 1
                    model_triggering_fetch = True
            
            if model_triggering_fetch:
                http_models_checked_count += 1

    if is_capped:
        issues.append(ExportValidationIssue(
            severity="warning",
            message=f"HTTP image checks capped. Checked {http_urls_scheduled_count} new URLs across {http_models_checked_count} models (cap reached)."
        ))

    # Execute New HTTP Checks
    urls_to_fetch = list(set(job['url'] for job in verification_jobs if not job['cached']))
    
    if urls_to_fetch:
        def check_url_status(u):
            try:
                # Short timeout, use HEAD to be lightweight
                r = requests.head(u, timeout=2.0, allow_redirects=True)
                return u, r.status_code, None
            except Exception as e:
                return u, None, str(e)
        
        fresh_results = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_CONCURRENCY) as executor:
             future_to_url = {executor.submit(check_url_status, u): u for u in urls_to_fetch}
             for future in concurrent.futures.as_completed(future_to_url):
                 u, code, err = future.result()
                 fresh_results[u] = (code, err)
                 
    # Update Cache and Generate Issues
    now = time.time()
    
    for job in verification_jobs:
        url = job['url']
        code = None
        err = None
        
        if job['cached']:
            code = job['entry']['code']
            err = job['entry']['error']
        else:
             if url in fresh_results:
                 code, err = fresh_results[url]
                 # Update Cache
                 HTTP_CACHE[url] = {'time': now, 'code': code, 'error': err}
        
        if err:
             issues.append(ExportValidationIssue(
                 severity="warning",
                 model_id=job['model'].id,
                 model_name=job['model'].name,
                 message=f"Image URL inaccessible ({err}): {url}"
                 ))
        elif code and code >= 400:
             issues.append(ExportValidationIssue(
                 severity="warning", 
                 model_id=job['model'].id, 
                 model_name=job['model'].name, 
                 message=f"Image URL inaccessible (HTTP {code}): {url}"
            ))

    # Determine Status
    error_count = sum(1 for i in issues if i.severity == "error")
    warning_count = sum(1 for i in issues if i.severity == "warning")
    
    status = "valid"
    if error_count > 0:
        status = "errors"
    elif warning_count > 0:
        status = "warnings"
        
    return ExportValidationResponse(
        status=status,
        summary_counts={
            "total_models": total_models,
            "issues": len(issues),
            "errors": error_count,
            "warnings": warning_count
        },
        items=issues
    )

@router.post("/preview", response_model=ExportPreviewResponse)
def generate_export_preview(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    recalc_performed, recalc_count = ensure_models_fresh_for_export(request, db)

    header_rows, data_rows, filename_base, template_code, models, ordered_field_names, ordered_required_flags = build_export_data(request, db)
    
    signature = compute_export_signature(template_code, header_rows, data_rows)
    field_map = {name: idx for idx, name in enumerate(ordered_field_names)}
    
    # Audit Construction
    audit_columns = build_audit_columns(ordered_field_names, ordered_required_flags)
    matched_price_fields = []
    matched_sku_fields = []
    
    for col in audit_columns:
        if col["type"] == "PRICE":
            matched_price_fields.append({
                "field_name": col["field_name"],
                "column_index": col["column_index"],
                "rule_explanation": (
                    "Price is populated from the Amazon US baseline pricing snapshot "
                    "(variant: choice_no_padding). Export fails if snapshot is missing."
                ),
                "source": {
                    "marketplace": "amazon",
                    "variant_key": "choice_no_padding",
                    "entity": "ModelPricingSnapshot",
                    "field": "retail_price_cents"
                }
            })
        elif col["type"] == "SKU":
             explanation = "Value is derived from the model base SKU."
             if "contribution_sku" in col["field_name"].lower() and request.listing_type == "individual":
                 explanation += " In individual listings, contribution_sku uses the base/parent SKU."
                 
             matched_sku_fields.append({
                "field_name": col["field_name"],
                "column_index": col["column_index"],
                "rule_explanation": explanation
            })

    row_samples = []
    for i in range(min(5, len(data_rows))):
        row = data_rows[i]
        model = models[i]
        
        key_values = {}
        for col in audit_columns:
            # Safely get value if index is within bounds, though it should be
            idx = col["column_index"]
            if idx < len(row):
                key_values[col["field_name"]] = row[idx]
        
        row_samples.append({
            "model_id": model.id,
            "model_name": model.name,
            "key_values": key_values
        })

    # Field Source Audit (for first model)
    field_sources = []
    if models:
        audit_model = models[0]
        audit_series = db.query(Series).filter(Series.id == audit_model.series_id).first()
        audit_mfr = db.query(Manufacturer).filter(Manufacturer.id == audit_series.manufacturer_id).first() if audit_series else None
        audit_equip = db.query(EquipmentType).filter(EquipmentType.id == audit_model.equipment_type_id).first()
        
        link = db.query(EquipmentTypeProductType).filter(EquipmentTypeProductType.equipment_type_id == audit_model.equipment_type_id).first()
        if link:
            audit_fields = db.query(ProductTypeField).filter(ProductTypeField.product_type_id == link.product_type_id).all()
            for f in audit_fields:
               src = get_field_source_audit(f, audit_model, audit_series, audit_mfr, audit_equip, is_image_url_field(f.field_name))
               if src:
                   field_sources.append(src)

    audit = {
        "row_mode": "BASE",
        "pricing": {
            "matched_price_fields": matched_price_fields
        },
        "sku": {
            "matched_sku_fields": matched_sku_fields
        },
        "row_samples": row_samples,
        "field_sources": field_sources
    }
    
    rows = []
    # Reconstruct ExportRowData objects from the raw data and model objects
    # We rely on the order of 'models' and 'data_rows' matching, which they do in build_export_data
    for i, row_data in enumerate(data_rows):
        rows.append(ExportRowData(
            model_id=models[i].id,
            model_name=models[i].name,
            data=row_data
        ))
    
    return ExportPreviewResponse(
        headers=header_rows,
        rows=rows,
        template_code=template_code,
        export_signature=signature,
        field_map=field_map,
        audit=audit,
        recalc_performed=recalc_performed,
        recalc_model_count=recalc_count
    )


from datetime import datetime

def build_export_data(request: ExportPreviewRequest, db: Session):
    """Build export data (headers and rows) for the given models."""
    if not request.model_ids:
        raise HTTPException(status_code=400, detail="No models selected")
    
    models = db.query(Model).filter(Model.id.in_(request.model_ids)).all()
    if not models:
        raise HTTPException(status_code=404, detail="No models found")
    
    equipment_type_ids = set(m.equipment_type_id for m in models)
    if len(equipment_type_ids) > 1:
        raise HTTPException(
            status_code=400, 
            detail="All selected models must have the same equipment type for export"
        )
    
    equipment_type_id = list(equipment_type_ids)[0]
    
    link = db.query(EquipmentTypeProductType).filter(
        EquipmentTypeProductType.equipment_type_id == equipment_type_id
    ).first()
    
    if not link:
        equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
        raise HTTPException(
            status_code=400, 
            detail=f"No Amazon template linked to equipment type: {equipment_type.name if equipment_type else 'Unknown'}"
        )
    
    product_type = db.query(AmazonProductType).filter(
        AmazonProductType.id == link.product_type_id
    ).first()
    
    if not product_type:
        raise HTTPException(status_code=404, detail="Template not found")
    
    fields = db.query(ProductTypeField).filter(
        ProductTypeField.product_type_id == product_type.id
    ).order_by(ProductTypeField.order_index).all()
    
    header_rows = product_type.header_rows or []
    
    equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
    
    first_model = models[0]
    first_series = db.query(Series).filter(Series.id == first_model.series_id).first()
    first_manufacturer = db.query(Manufacturer).filter(Manufacturer.id == first_series.manufacturer_id).first() if first_series else None
    
    mfr_name = normalize_for_url(first_manufacturer.name) if first_manufacturer else 'Unknown'
    series_name = normalize_for_url(first_series.name) if first_series else 'Unknown'
    date_str = datetime.now().strftime('%Y%m%d')
    filename_base = f"Amazon_{mfr_name}_{series_name}_{date_str}"
    
    data_rows = []
    for model in models:
        series = db.query(Series).filter(Series.id == model.series_id).first()
        manufacturer = db.query(Manufacturer).filter(Manufacturer.id == series.manufacturer_id).first() if series else None
        
        row_data = []
        for field in fields:
            if not field.required:
                row_data.append("")
                continue

            value = get_field_value(field, model, series, manufacturer, equipment_type, request.listing_type, db)
            row_data.append(value if value else '')
        
        data_rows.append(row_data)
    
    ordered_field_names = [f.field_name for f in fields]
    ordered_required_flags = [f.required for f in fields]
    return header_rows, data_rows, filename_base, product_type.code, models, ordered_field_names, ordered_required_flags


@router.post("/download/xlsx")
def download_xlsx(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """Download export as XLSX file."""
    ensure_models_fresh_for_export(request, db)
    header_rows, data_rows, filename_base, template_code, _, _, _ = build_export_data(request, db)
    
    sig = compute_export_signature(template_code, header_rows, data_rows)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    row_styles = [
        (Font(bold=True, color="FFFFFF"), PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")),
        (Font(color="FFFFFF"), PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")),
        (Font(bold=True, color="FFFFFF"), PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")),
        (Font(bold=True), PatternFill(start_color="8BC34A", end_color="8BC34A", fill_type="solid")),
        (Font(size=9), PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")),
        (Font(italic=True, size=9), PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")),
    ]
    
    current_row = 1
    for row_idx, header_row in enumerate(header_rows):
        for col_idx, value in enumerate(header_row):
            cell = ws.cell(row=current_row, column=col_idx + 1, value=value or '')
            if row_idx < len(row_styles):
                cell.font = row_styles[row_idx][0]
                cell.fill = row_styles[row_idx][1]
            cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
    
    for data_row in data_rows:
        for col_idx, value in enumerate(data_row):
            ws.cell(row=current_row, column=col_idx + 1, value=value or '')
        current_row += 1
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = min(len(str(cell.value)), 50)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{filename_base}.xlsx"
    response = StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
    response.headers["X-Export-Signature"] = sig
    response.headers["X-Export-Template-Code"] = template_code
    return response


@router.post("/download/xlsm")
def download_xlsm(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """Download export as XLSM file (macro-enabled workbook)."""
    ensure_models_fresh_for_export(request, db)
    header_rows, data_rows, filename_base, template_code, _, _, _ = build_export_data(request, db)

    sig = compute_export_signature(template_code, header_rows, data_rows)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    row_styles = [
        (Font(bold=True, color="FFFFFF"), PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")),
        (Font(color="FFFFFF"), PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")),
        (Font(bold=True, color="FFFFFF"), PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")),
        (Font(bold=True), PatternFill(start_color="8BC34A", end_color="8BC34A", fill_type="solid")),
        (Font(size=9), PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")),
        (Font(italic=True, size=9), PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")),
    ]
    
    current_row = 1
    for row_idx, header_row in enumerate(header_rows):
        for col_idx, value in enumerate(header_row):
            cell = ws.cell(row=current_row, column=col_idx + 1, value=value or '')
            if row_idx < len(row_styles):
                cell.font = row_styles[row_idx][0]
                cell.fill = row_styles[row_idx][1]
            cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
    
    for data_row in data_rows:
        for col_idx, value in enumerate(data_row):
            ws.cell(row=current_row, column=col_idx + 1, value=value or '')
        current_row += 1
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = min(len(str(cell.value)), 50)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{filename_base}.xlsm"
    response = StreamingResponse(
        output,
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
    response.headers["X-Export-Signature"] = sig
    response.headers["X-Export-Template-Code"] = template_code
    return response


@router.post("/download/csv")
def download_csv(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """Download export as CSV file."""
    ensure_models_fresh_for_export(request, db)
    header_rows, data_rows, filename_base, template_code, _, _, _ = build_export_data(request, db)
    
    sig = compute_export_signature(template_code, header_rows, data_rows)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    for header_row in header_rows:
        writer.writerow([v or '' for v in header_row])
    
    for data_row in data_rows:
        writer.writerow([v or '' for v in data_row])
    
    content = output.getvalue().encode('utf-8')
    
    filename = f"{filename_base}.csv"
    response = StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
    response.headers["X-Export-Signature"] = sig
    response.headers["X-Export-Template-Code"] = template_code
    return response



def generate_customization_unicode_txt(template_path: str, skus: List[str]) -> bytes:
    """
    Generate Amazon Customization File (.txt) from a template and list of SKUs.
    Format:
    - UTF-16 LE with BOM
    - Tab-delimited
    - Rows 1-3: from template unmodified
    - Row 4: Blueprint row (from template)
    - Rows 5+: Blueprint row with Column A replaced by real SKU
    """
    try:
        wb = load_workbook(template_path, read_only=False, data_only=True)
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read customization template: {str(e)}")

    if len(rows) < 4:
         raise HTTPException(status_code=500, detail="Customization template is invalid (fewer than 4 rows).")

    # Capture Headers (1-3) and Blueprint (4)
    header_rows = rows[:3]
    blueprint_row = rows[3]
    
    output = io.BytesIO()
    # Write BOM for UTF-16 LE
    output.write(b'\xff\xfe')
    
    def write_row(row_values):
        # Normalize: 
        # - None -> ""
        # - strip tabs/newlines from content to prevent breaking format
        # - join with tabs
        cleaned = []
        for val in row_values:
            s = str(val) if val is not None else ""
            s = s.replace('\t', ' ').replace('\r', ' ').replace('\n', ' ')
            cleaned.append(s)
        
        line = "\t".join(cleaned) + "\r\n"
        output.write(line.encode('utf-16-le'))

    # Write Headers
    for r in header_rows:
        write_row(r)
        
    # Write Data Rows
    for sku in skus:
        # Create new row based on blueprint
        new_row = list(blueprint_row)
        # Column A is index 0
        if len(new_row) > 0:
            new_row[0] = sku
        else:
            new_row = [sku]
            
        write_row(new_row)
        
    return output.getvalue()


class DownloadZipRequest(BaseModel):
    model_ids: List[int]
    listing_type: str = "individual"
    include_customization: bool = True
    marketplace_token: str
    manufacturer_token: str
    series_token: str
    series_token: str
    date_token: str  # YYYY-MM-DD
    customization_format: Optional[str] = "xlsx" # "xlsx" or "txt"

@router.post("/download/zip")
def download_zip(request: DownloadZipRequest, db: Session = Depends(get_db)):
    """
    Download a ZIP package containing:
    1. XLSM (Macro-Enabled)
    2. XLSX (Standard)
    3. CSV (Data Only)
    Optional: 4. Customization .txt (if toggle ON and rules apply - future chunk)
    
    Filenames are strictly constructed from tokens provided by the UI.
    """
    # 1. Validation & Data Build (Reuse existing logic)
    # We map DownloadZipRequest to ExportPreviewRequest for the helper
    preview_req = ExportPreviewRequest(model_ids=request.model_ids, listing_type=request.listing_type)
    ensure_models_fresh_for_export(preview_req, db)
    
    header_rows, data_rows, _, template_code, exported_models, ordered_field_names, _ = build_export_data(preview_req, db)
    
    # 2. Compute Filename Base
    # Syntax: [Marketplace]-[Manufacturer]-[Series]-Product_Upload-[Date]
    filename_base = f"{request.marketplace_token}-{request.manufacturer_token}-{request.series_token}-Product_Upload-{request.date_token}"
    
    # 3. Generate Files In-Memory
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        
        # --- A. Generate XLSX ---
        wb_xlsx = Workbook()
        ws_xlsx = wb_xlsx.active
        ws_xlsx.title = "Template"
        
        # Styles (Same as download_xlsx)
        row_styles = [
            (Font(bold=True, color="FFFFFF"), PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")),
            (Font(color="FFFFFF"), PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")),
            (Font(bold=True, color="FFFFFF"), PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")),
            (Font(bold=True), PatternFill(start_color="8BC34A", end_color="8BC34A", fill_type="solid")),
            (Font(size=9), PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")),
            (Font(italic=True, size=9), PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")),
        ]
        
        current_row = 1
        for row_idx, header_row in enumerate(header_rows):
            for col_idx, value in enumerate(header_row):
                cell = ws_xlsx.cell(row=current_row, column=col_idx + 1, value=value or '')
                if row_idx < len(row_styles):
                    cell.font = row_styles[row_idx][0]
                    cell.fill = row_styles[row_idx][1]
                cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1
        
        for data_row in data_rows:
            for col_idx, value in enumerate(data_row):
                ws_xlsx.cell(row=current_row, column=col_idx + 1, value=value or '')
            current_row += 1
            
        # Column Widths
        for col in ws_xlsx.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)
                except:
                    pass
            ws_xlsx.column_dimensions[column].width = max_length + 2

        xlsx_buffer = io.BytesIO()
        wb_xlsx.save(xlsx_buffer)
        zf.writestr(f"{filename_base}.xlsx", xlsx_buffer.getvalue())
        
        # --- B. Generate XLSM ---
        # Note: logic is identical to XLSX for openpyxl, but MIME/vbaProject handling might differ in real apps.
        # openpyxl saves as macro enabled if asked, but doesn't add macros dynamically.
        # existing download_xlsm endpoint essentially saves standard xlsx as .xlsm MIME.
        # We replicate that exact behavior here.
        wb_xlsm = Workbook()
        ws_xlsm = wb_xlsm.active
        ws_xlsm.title = "Template"
        
        current_row = 1
        for row_idx, header_row in enumerate(header_rows):
            for col_idx, value in enumerate(header_row):
                cell = ws_xlsm.cell(row=current_row, column=col_idx + 1, value=value or '')
                if row_idx < len(row_styles):
                    cell.font = row_styles[row_idx][0]
                    cell.fill = row_styles[row_idx][1]
                cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1
        
        for data_row in data_rows:
            for col_idx, value in enumerate(data_row):
                ws_xlsm.cell(row=current_row, column=col_idx + 1, value=value or '')
            current_row += 1
            
        for col in ws_xlsm.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)
                except:
                    pass
            ws_xlsm.column_dimensions[column].width = max_length + 2
            
        xlsm_buffer = io.BytesIO()
        wb_xlsm.save(xlsm_buffer)
        zf.writestr(f"{filename_base}.xlsm", xlsm_buffer.getvalue())

        # --- C. Generate CSV ---
        csv_output = io.StringIO()
        writer = csv.writer(csv_output)
        for header_row in header_rows:
            writer.writerow([v or '' for v in header_row])
        for data_row in data_rows:
            writer.writerow([v or '' for v in data_row])
        
        zf.writestr(f"{filename_base}.csv", csv_output.getvalue().encode('utf-8'))
        
        # --- D. Customization File ---
        if request.include_customization:
            # 1. Identify SKU column from export data
            sku_col_idx = -1
            target_fields = ["item_sku", "contribution_sku", "external_product_id"] # Fallbacks
            
            for target in target_fields:
                for idx, field_name in enumerate(ordered_field_names):
                    if target in field_name.lower():
                        sku_col_idx = idx
                        break
                if sku_col_idx != -1:
                    break
            
            # If still found, fallback to Model.parent_sku? 
            # Requirement says "skus[] from the Amazon worksheet export rows (authoritative)"
            # If not found in export, we likely cannot proceed safely with an authoritative matching.
            # But let's check if we can fallback to index 0.
            if sku_col_idx == -1:
                # Fallback: Assume first column is SKU if named appropriately? Or just use index 0.
                # Common Amazon template practice: Column A is SKU.
                if ordered_field_names and "sku" in ordered_field_names[0].lower():
                     sku_col_idx = 0
            
            skus = []
            if sku_col_idx != -1:
                for row in data_rows:
                    if sku_col_idx < len(row):
                         skus.append(row[sku_col_idx])
                    else:
                         skus.append("")
            else:
                 # Fallback to model parent_sku if we can't find it in the sheet?
                 # This violates "authoritative from export rows", but is safer than crashing.
                 # Actually, let's log/warn and skip if we can't find SKUs.
                 pass

            # Audit / Reconciliation Check
            worksheet_count = len(data_rows)
            skus_count = len(skus)
            
            # Robustness: Check for duplicates and content signature
            unique_skus = set(skus)
            dupe_count = skus_count - len(unique_skus)
            
            # Deterministic Signature (SHA256 of joined SKUs)
            # Normalize to empty string if None before joining, though skus list extraction handled that above somewhat. (appended "")
            # Ensure all are strings
            clean_skus_for_sig = [str(s) if s is not None else "" for s in skus]
            sku_signature = hashlib.sha256("\n".join(clean_skus_for_sig).encode("utf-8")).hexdigest()[:12]
            
            first_sku = clean_skus_for_sig[0] if clean_skus_for_sig else "<none>"
            last_sku = clean_skus_for_sig[-1] if clean_skus_for_sig else "<none>"
            
            # Log full audit
            print(f"[EXPORT] SKU Audit: signature={sku_signature} count={skus_count} dupes={dupe_count} first={first_sku} last={last_sku}")
            
            should_skip = False
            if skus_count != worksheet_count:
                print(f"[EXPORT][ERROR] Customization SKU count mismatch (rows={worksheet_count}, skus={skus_count}).")
                should_skip = True
            
            if dupe_count > 0:
                print(f"[EXPORT][ERROR] Duplicate SKUs detected in export ({dupe_count}). Customization template requires unique SKUs.")
                should_skip = True
                
            if should_skip:
                print(f"[EXPORT][ERROR] Skipping customization file generation due to audit failure.")
                skus = [] # Force skip

            if skus:
                 # 2. Determine Template
                 # Group models by equipment type to find template.
                 # build_export_data enforces single equipment_type for the batch.
                 # So we just check the first model.
                 if exported_models:
                     first_model = exported_models[0]
                     equip_type = db.query(EquipmentType).filter(EquipmentType.id == first_model.equipment_type_id).first()
                     
                     if equip_type and equip_type.amazon_customization_template_id:
                         template = equip_type.amazon_customization_template
                         # We need the full path
                         
                         if os.path.exists(template.file_path):
                             try:
                                 # Format Logic:
                                 fmt = (request.customization_format or "xlsx").lower()

                                 if fmt == 'xlsx':
                                     # --- XLSX Mode ---
                                     # Read raw template file and write to ZIP (byte-identical preservation)
                                     print(f"[EXPORT] Customization XLSX: template_id={template.id} path={template.file_path}")

                                     if not template.file_path:
                                         raise HTTPException(status_code=500, detail="Assigned customization template has no file_path")

                                     if not os.path.exists(template.file_path):
                                         raise HTTPException(status_code=500, detail=f"Assigned customization template file missing: {template.file_path}")

                                     with open(template.file_path, "rb") as f:
                                         cust_bytes = f.read()

                                     if not cust_bytes:
                                         raise HTTPException(status_code=500, detail=f"Assigned customization template is empty on disk: {template.file_path}")
                                     
                                     cust_filename = f"{request.marketplace_token}-{request.manufacturer_token}-{request.series_token}-Customization-{request.date_token}.xlsx"
                                     zf.writestr(cust_filename, cust_bytes)
                                 
                                 else:
                                     # --- TXT Mode (Legacy) ---
                                     cust_bytes = generate_customization_unicode_txt(template.file_path, skus)
                                     cust_filename = f"{request.marketplace_token}-{request.manufacturer_token}-{request.series_token}-Customization-{request.date_token}.txt"
                                     zf.writestr(cust_filename, cust_bytes)
                                 
                             except Exception as e:
                                 # Log error but don't fail entire export? 
                                 # User said "Hard error" if multiple templates. Here it's generation failure.
                                 # Failing safe for now by raising to alert user.
                                 raise HTTPException(status_code=500, detail=f"Failed to generate customization file: {str(e)}")
                         else:
                             # File missing on disk
                             raise HTTPException(status_code=500, detail=f"Assigned customization template file missing: {template.file_path}")
                     # Else: No assignment -> Skip (as per requirements)
            
    # Helper for customization generation
    # ... logic ...
    
# Wait, I need to do this in steps.
# 1. Update imports.
# 2. Update a helper function `generate_customization_unicode_txt`.
# 3. Update `download_zip` to capture `ordered_field_names` and implement the logic.

# Let's do imports first.

            
    zip_buffer.seek(0)
    
    zip_filename = f"{filename_base}.zip"
    
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={zip_filename}"}
    )

IMAGE_FIELD_TO_NUMBER = {

    'main_product_image_locator': '001',
    'other_product_image_locator_1': '002',
    'other_product_image_locator_2': '003',
    'other_product_image_locator_3': '004',
    'other_product_image_locator_4': '005',
    'other_product_image_locator_5': '006',
    'other_product_image_locator_6': '007',
    'other_product_image_locator_7': '008',
    'other_product_image_locator_8': '009',
    'swatch_product_image_locator': '010',
}

def normalize_for_url(name: str) -> str:
    """Normalize a name for use in URL paths/filenames.
    Removes spaces, special characters, and non-alphanumeric characters.
    Example: "Fender USA" -> "FenderUSA", "Tone-Master" -> "ToneMaster"
    """
    if not name:
        return ''
    result = re.sub(r'[^a-zA-Z0-9]', '', name)
    return result

def substitute_placeholders(value: str, model: Model, series, manufacturer, equipment_type, is_image_url: bool = False) -> str:
    result = value
    mfr_name = manufacturer.name if manufacturer else ''
    series_name = series.name if series else ''
    model_name = model.name if model else ''
    equip_type = equipment_type.name if equipment_type else ''
    
    if is_image_url:
        def img_norm(s):
            if not s: return ''
            # Replace whitespace sequences with underscore, keep other chars including punctuation
            return re.sub(r'\s+', '_', s.strip())

        mfr_name_norm = img_norm(mfr_name)
        series_name_norm = img_norm(series_name)
        model_name_norm = img_norm(model_name)
        equip_type_norm = img_norm(equip_type)
        
        result = result.replace('[Manufacturer_Name]', mfr_name_norm)
        result = result.replace('[Series_Name]', series_name_norm)
        result = result.replace('[Model_Name]', model_name_norm)
        result = result.replace('[MANUFACTURER_NAME]', mfr_name_norm)
        result = result.replace('[SERIES_NAME]', series_name_norm)
        result = result.replace('[MODEL_NAME]', model_name_norm)
        
        result = result.replace('[EQUIPMENT_TYPE]', equip_type_norm)
        result = result.replace('[Equipment_Type]', equip_type_norm)
    else:
        result = result.replace('[MANUFACTURER_NAME]', mfr_name)
        result = result.replace('[SERIES_NAME]', series_name)
        result = result.replace('[MODEL_NAME]', model_name)
        result = result.replace('[Manufacturer_Name]', mfr_name)
        result = result.replace('[Series_Name]', series_name)
        result = result.replace('[Model_Name]', model_name)
    
        result = result.replace('[EQUIPMENT_TYPE]', equip_type)
        result = result.replace('[Equipment_Type]', equip_type)
    
    return result

def get_image_field_key(field_name: str) -> str | None:
    """Extract the base image field key from a full Amazon field name.
    Returns the key if it matches a known product image field, None otherwise.
    """
    for key in IMAGE_FIELD_TO_NUMBER.keys():
        if field_name.startswith(key):
            return key
    return None

def is_image_url_field(field_name: str) -> bool:
    """Check if a field is a product image URL field that needs special processing."""
    return get_image_field_key(field_name) is not None

def get_amazon_us_baseline_price_str(db: Session, model_id: int) -> str:
    """
    Fetch the baseline retail price (Choice Waterproof No Padding) for Amazon US.
    Format: "249.95" (2 decimals)
    Strict Rule: Fail if snapshot is missing.
    """
    snapshot = db.query(ModelPricingSnapshot).filter(
        ModelPricingSnapshot.model_id == model_id,
        ModelPricingSnapshot.marketplace == "amazon",
        ModelPricingSnapshot.variant_key == "choice_no_padding"
    ).first()

    if not snapshot:
        # Prompt: "If baseline snapshot row is missing, fail the export for that model with a clear message"
        # Since this is deep in the call stack for a specific field, raising HTTPException here
        # will abort the entire request, which is desired.
        raise HTTPException(
            status_code=400, 
            detail=f"Missing baseline pricing snapshot for Choice Waterproof (no padding). Run pricing recalculation for model {model_id} on 'amazon' marketplace before exporting."
        )
    
    return f"{snapshot.retail_price_cents / 100:.2f}"

def get_field_value(field: ProductTypeField, model: Model, series, manufacturer, equipment_type=None, listing_type: str = "individual", db: Session = None) -> str | None:
    field_name_lower = field.field_name.lower()
    is_image_field = is_image_url_field(field.field_name)
    
    # Phase 9: Amazon Baseline Price Logic
    # check specific field key parts
    if "purchasable_offer[marketplace_id=atvpdkikx0der]" in field_name_lower and "our_price#1.schedule#1.value_with_tax" in field_name_lower:
        if db:
            return get_amazon_us_baseline_price_str(db, model.id)
        # If db is somehow missing (shouldn't happen with updated callers), strict rules say NO FALLBACK.
        # But for now, if db is missing, we can't query.
        raise HTTPException(status_code=500, detail="Database session missing in export logic.")
    
    if 'contribution_sku' in field_name_lower and listing_type == 'individual':
        return model.parent_sku if model.parent_sku else None
    
    # Only include custom_value or selected_value if field is marked as required
    if field.required:
        if field.selected_value:
            return substitute_placeholders(field.selected_value, model, series, manufacturer, equipment_type, is_image_url=is_image_field)
            
        if field.custom_value:
            return substitute_placeholders(field.custom_value, model, series, manufacturer, equipment_type, is_image_url=is_image_field)
        
        # Auto-generate values for common fields only if required
        if 'item_name' in field_name_lower or 'product_name' in field_name_lower or 'title' in field_name_lower:
            mfr_name = manufacturer.name if manufacturer else ''
            series_name = series.name if series else ''
            return f"{mfr_name} {series_name} {model.name} Cover"
        
        if 'brand' in field_name_lower or 'brand_name' in field_name_lower:
            return manufacturer.name if manufacturer else None
        
        if 'model' in field_name_lower or 'model_number' in field_name_lower or 'model_name' in field_name_lower:
            return model.name
        
        if 'manufacturer' in field_name_lower:
            return manufacturer.name if manufacturer else None
    
    return None
