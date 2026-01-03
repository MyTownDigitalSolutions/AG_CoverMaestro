from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session, selectinload
from sqlalchemy import func
from typing import Optional, List
import os
from openpyxl import load_workbook

from app.database import get_db
from app.services.ebay_template_service import EbayTemplateService
from app.schemas.templates import (
    EbayTemplateResponse,
    EbayTemplateParseSummary,
    EbayTemplateFieldsResponse,
    EbayFieldResponse,
    EbayFieldUpdateRequest,
    EbayValidValueCreateRequest,
    EbayTemplatePreviewResponse
)
from app.models.templates import EbayTemplate, EbayField, EbayFieldValue

router = APIRouter(
    prefix="/ebay-templates",
    tags=["Ebay Templates"]
)


def _build_ebay_template_fields_response(template_id: int, db: Session) -> EbayTemplateFieldsResponse:
    """
    Internal helper: Load fields + valid values for a template and map to API response models.
    Uses selectinload to reliably populate one-to-many relationships without join edge-cases.
    """
    # 1) Verify template exists
    template = db.query(EbayTemplate).filter(EbayTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    # 2) Query fields and eagerly load valid_values via SELECT IN (more reliable than joinedload)
    # Order by order_index ASC (nulls last), then id ASC
    fields: List[EbayField] = (
        db.query(EbayField)
        .options(selectinload(EbayField.valid_values))
        .filter(EbayField.ebay_template_id == template_id)
        .order_by(func.coalesce(EbayField.order_index, 10**9), EbayField.id)
        .all()
    )

    # 3) Map to response
    response_fields: List[EbayFieldResponse] = []
    for f in fields:
        # Sort values deterministically by ID ASC
        sorted_values = sorted((f.valid_values or []), key=lambda v: v.id)

        # Map to List[str] as required by schema field 'allowed_values'
        allowed_strs = [v.value for v in sorted_values]
        
        # Map to detailed list with IDs for delete operations
        allowed_detailed = [{"id": v.id, "value": v.value} for v in sorted_values]

        response_fields.append(
            EbayFieldResponse(
                id=f.id,
                ebay_template_id=f.ebay_template_id,
                field_name=f.field_name,
                display_name=f.display_name,
                required=f.required,
                order_index=f.order_index,
                selected_value=f.selected_value,
                custom_value=f.custom_value,
                allowed_values=allowed_strs,
                allowed_values_detailed=allowed_detailed
            )
        )

    return EbayTemplateFieldsResponse(
        template_id=template_id,
        fields=response_fields
    )


@router.post("/upload", response_model=EbayTemplateResponse)
async def upload_ebay_template(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Upload and store the canonical eBay XLSX template (bit-for-bit).
    """
    service = EbayTemplateService(db)
    return await service.store_ebay_template_upload(file)


@router.get("/current", response_model=Optional[EbayTemplateResponse])
def get_current_ebay_template(db: Session = Depends(get_db)):
    """
    Get the most recently uploaded eBay template metadata.
    """
    latest = (
        db.query(EbayTemplate)
        .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
        .first()
    )

    if not latest:
        return None

    return latest


@router.post("/{template_id}/parse", response_model=EbayTemplateParseSummary)
def parse_ebay_template(
    template_id: int,
    db: Session = Depends(get_db)
):
    """
    Parse the eBay template file and populate metadata in the database.
    Idempotent operation (clears existing fields/values for this template).
    """
    service = EbayTemplateService(db)
    return service.parse_ebay_template(template_id)


@router.get("/current/fields", response_model=EbayTemplateFieldsResponse)
def get_current_ebay_template_fields(db: Session = Depends(get_db)):
    """
    Get the parsed fields and allowed values for the MOST RECENT template.
    """
    latest = (
        db.query(EbayTemplate)
        .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
        .first()
    )

    if not latest:
        raise HTTPException(status_code=404, detail="No eBay template uploaded")

    return _build_ebay_template_fields_response(latest.id, db)


@router.get("/{template_id}/fields", response_model=EbayTemplateFieldsResponse)
def get_ebay_template_fields(template_id: int, db: Session = Depends(get_db)):
    """
    Get the parsed fields and allowed values for a specific template.
    Returns fields ordered by order_index.
    """
    return _build_ebay_template_fields_response(template_id, db)


@router.patch("/fields/{field_id}", response_model=EbayFieldResponse)
def update_ebay_field(
    field_id: int,
    updates: EbayFieldUpdateRequest,
    db: Session = Depends(get_db)
):
    """
    Update eBay field properties (required, selected_value, custom_value).
    
    Validation rules:
    - selected_value "Any" → stored as None
    - selected_value must exist in valid_values OR custom_value must be set
    - empty custom_value → stored as None
    """
    # Load field with valid values for validation
    field = (
        db.query(EbayField)
        .options(selectinload(EbayField.valid_values))
        .filter(EbayField.id == field_id)
        .first()
    )
    
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    
    # Update required if provided
    if updates.required is not None:
        field.required = updates.required
    
    # Update selected_value if provided
    if updates.selected_value is not None:
        if updates.selected_value == "Any" or updates.selected_value == "":
            field.selected_value = None
        else:
            # Get list of valid values
            valid_values_list = [v.value for v in field.valid_values]
            
            # Allow if it's in valid values OR if custom_value is being set
            if updates.selected_value in valid_values_list or updates.custom_value is not None:
                field.selected_value = updates.selected_value
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid selected_value. Must be one of {valid_values_list} or set custom_value"
                )
    
    # Update custom_value if provided
    if updates.custom_value is not None:
        field.custom_value = updates.custom_value if updates.custom_value else None
    
    # Commit changes
    db.commit()
    db.refresh(field)
    
    # Return updated field using same response structure
    sorted_values = sorted((field.valid_values or []), key=lambda v: v.id)
    allowed_strs = [v.value for v in sorted_values]
    allowed_detailed = [{"id": v.id, "value": v.value} for v in sorted_values]
    
    return EbayFieldResponse(
        id=field.id,
        ebay_template_id=field.ebay_template_id,
        field_name=field.field_name,
        display_name=field.display_name,
        required=field.required,
        order_index=field.order_index,
        selected_value=field.selected_value,
        custom_value=field.custom_value,
        allowed_values=allowed_strs,
        allowed_values_detailed=allowed_detailed
    )


@router.post("/fields/{field_id}/valid-values", response_model=EbayFieldResponse)
def add_valid_value_to_field(
    field_id: int,
    request: EbayValidValueCreateRequest,
    db: Session = Depends(get_db)
):
    """
    Add a valid value to an eBay field.
    
    Rules:
    - Trims whitespace
    - Rejects empty values
    - De-duplicates (case-sensitive)
    - Returns updated field with all valid values
    """
    # Trim and validate
    value = request.value.strip()
    if not value:
        raise HTTPException(status_code=400, detail="Value cannot be empty")
    
    # Load field with valid values
    field = (
        db.query(EbayField)
        .options(selectinload(EbayField.valid_values))
        .filter(EbayField.id == field_id)
        .first()
    )
    
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    
    # Check if value already exists (case-sensitive)
    existing = (
        db.query(EbayFieldValue)
        .filter(
            EbayFieldValue.ebay_field_id == field_id,
            EbayFieldValue.value == value
        )
        .first()
    )
    
    if not existing:
        # Add new value
        new_value = EbayFieldValue(ebay_field_id=field_id, value=value)
        db.add(new_value)
        db.commit()
        db.refresh(field)
    
    # Return updated field
    sorted_values = sorted((field.valid_values or []), key=lambda v: v.id)
    allowed_strs = [v.value for v in sorted_values]
    allowed_detailed = [{"id": v.id, "value": v.value} for v in sorted_values]
    
    return EbayFieldResponse(
        id=field.id,
        ebay_template_id=field.ebay_template_id,
        field_name=field.field_name,
        display_name=field.display_name,
        required=field.required,
        order_index=field.order_index,
        selected_value=field.selected_value,
        custom_value=field.custom_value,
        allowed_values=allowed_strs,
        allowed_values_detailed=allowed_detailed
    )


@router.delete("/fields/{field_id}/valid-values/{value_id}", response_model=EbayFieldResponse)
def delete_valid_value_from_field(
    field_id: int,
    value_id: int,
    db: Session = Depends(get_db)
):
    """
    Delete a valid value from an eBay field.
    
    Rules:
    - Ensures value exists and belongs to the field
    - If deleted value equals field.selected_value, clears selected_value
    - Returns updated field
    """
    # Load the value and verify it belongs to this field
    value_obj = (
        db.query(EbayFieldValue)
        .filter(
            EbayFieldValue.id == value_id,
            EbayFieldValue.ebay_field_id == field_id
        )
        .first()
    )
    
    if not value_obj:
        raise HTTPException(
            status_code=404,
            detail="Valid value not found or does not belong to this field"
        )
    
    # Load field
    field = (
        db.query(EbayField)
        .options(selectinload(EbayField.valid_values))
        .filter(EbayField.id == field_id)
        .first()
    )
    
    if not field:
        raise HTTPException(status_code=404, detail="Field not found")
    
    # If this value is currently selected, clear selected_value
    if field.selected_value == value_obj.value:
        field.selected_value = None
    
    # Delete the value
    db.delete(value_obj)
    db.commit()
    db.refresh(field)
    
    # Return updated field
    sorted_values = sorted((field.valid_values or []), key=lambda v: v.id)
    allowed_strs = [v.value for v in sorted_values]
    allowed_detailed = [{"id": v.id, "value": v.value} for v in sorted_values]
    
    return EbayFieldResponse(
        id=field.id,
        ebay_template_id=field.ebay_template_id,
        field_name=field.field_name,
        display_name=field.display_name,
        required=field.required,
        order_index=field.order_index,
        selected_value=field.selected_value,
        custom_value=field.custom_value,
        allowed_values=allowed_strs,
        allowed_values_detailed=allowed_detailed
    )


@router.get("/current/download")
def download_current_ebay_template(mode: str = "inline", db: Session = Depends(get_db)):
    """
    Download the current (latest) eBay template file as-is from disk.
    Returns the bit-for-bit original uploaded file.
    
    Args:
        mode: "inline" (try to display in browser) or "download" (force download)
    """
    # Get current template (newest by uploaded_at, then id desc)
    template = (
        db.query(EbayTemplate)
        .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
        .first()
    )
    
    if not template:
        raise HTTPException(status_code=404, detail="No eBay template found")
    
    # Check file exists on disk
    if not os.path.exists(template.file_path):
        raise HTTPException(status_code=400, detail="Template file not found on disk")
    
    # Set Content-Disposition based on mode
    disposition = "inline" if mode == "inline" else "attachment"
    
    # Return file
    response = FileResponse(
        path=template.file_path,
        filename=template.original_filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = f'{disposition}; filename="{template.original_filename}"'
    return response


@router.get("/current/preview", response_model=EbayTemplatePreviewResponse)
def preview_current_ebay_template(
    preview_rows: int = 25,
    preview_cols: int = 20,
    db: Session = Depends(get_db)
):
    """
    Preview the current (latest) eBay template as a grid.
    Returns a JSON grid of cell values for the top-left window.
    """
    # Get current template
    template = (
        db.query(EbayTemplate)
        .order_by(EbayTemplate.uploaded_at.desc(), EbayTemplate.id.desc())
        .first()
    )
    
    if not template:
        raise HTTPException(status_code=404, detail="No eBay template found")
    
    # Check file exists on disk
    if not os.path.exists(template.file_path):
        raise HTTPException(status_code=400, detail="Template file not found on disk")
    
    # Load workbook
    try:
        wb = load_workbook(template.file_path, data_only=True)
        
        # Choose sheet: prefer "Template", else first sheet
        if "Template" in wb.sheetnames:
            sheet = wb["Template"]
            sheet_name = "Template"
        else:
            sheet = wb.active
            sheet_name = sheet.title if sheet else "Unknown"
        
        # Get dimensions
        max_row = sheet.max_row or 0
        max_column = sheet.max_column or 0
        
        # Build grid for preview window
        grid = []
        for row_idx in range(1, min(preview_rows + 1, max_row + 1)):
            row_data = []
            for col_idx in range(1, min(preview_cols + 1, max_column + 1)):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # Convert to string safely
                if value is None:
                    row_data.append("")
                elif isinstance(value, (int, float)):
                    row_data.append(str(value))
                else:
                    row_data.append(str(value).strip())
            grid.append(row_data)
        
        wb.close()
        
        return EbayTemplatePreviewResponse(
            template_id=template.id,
            original_filename=template.original_filename,
            sheet_name=sheet_name,
            max_row=max_row,
            max_column=max_column,
            preview_row_count=len(grid),
            preview_column_count=len(grid[0]) if grid else 0,
            grid=grid
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load template: {str(e)}")
