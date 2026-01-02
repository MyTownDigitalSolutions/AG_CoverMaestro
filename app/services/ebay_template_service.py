import os
import hashlib
from fastapi import UploadFile, HTTPException
from sqlalchemy.orm import Session
from datetime import datetime
from typing import Optional
import openpyxl
import re

from app.models.templates import EbayTemplate, EbayField
# IMPORTANT: valid values model name differs in some codebases.
# We'll resolve it safely below.
from app.schemas.templates import EbayTemplateParseSummary
from app.services.storage_policy import (
    ensure_storage_dirs_exist,
    assert_allowed_write_path,
    get_ebay_template_paths,
    rotate_ebay_template_backup,
)


def _resolve_ebay_valid_value_model():
    """
    Resolve the ORM model used for EbayField.valid_values.
    Different codebases name it differently.
    We try common names in a safe order.
    """
    # Late imports so this file can still load even if a name doesn't exist.
    from app.models import templates as templates_models

    # Most common names (try in order)
    for name in ("EbayFieldValue", "EbayFieldValidValue", "EbayValidValue"):
        if hasattr(templates_models, name):
            return getattr(templates_models, name)

    raise RuntimeError(
        "Could not find a valid-values ORM model. Expected one of: "
        "EbayFieldValue, EbayFieldValidValue, EbayValidValue in app.models.templates"
    )


class EbayTemplateService:
    def __init__(self, db: Session):
        self.db = db
        self.ValidValueModel = _resolve_ebay_valid_value_model()

    async def store_ebay_template_upload(self, file: UploadFile) -> EbayTemplate:
        """
        Store an eBay template upload bit-for-bit.

        Steps:
        1. Read bytes & compute SHA256 upload
        2. Rotate existing file to backup
        3. Write new file to canonical path
        4. Re-read back from disk & compute SHA256 persisted
        5. Verify match
        6. Create DB record
        """
        if not file.filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

        ensure_storage_dirs_exist()
        canonical_path, backup_path = get_ebay_template_paths()

        assert_allowed_write_path(canonical_path)
        assert_allowed_write_path(backup_path)

        rotate_ebay_template_backup(canonical_path, backup_path)

        await file.seek(0)
        uploaded_bytes = await file.read()
        uploaded_sha256 = hashlib.sha256(uploaded_bytes).hexdigest()
        file_size = len(uploaded_bytes)

        try:
            with open(canonical_path, "wb") as f:
                f.write(uploaded_bytes)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to write file to disk: {e}")

        try:
            with open(canonical_path, "rb") as f:
                persisted_bytes = f.read()
            persisted_sha256 = hashlib.sha256(persisted_bytes).hexdigest()

            if uploaded_sha256 != persisted_sha256:
                if os.path.exists(canonical_path):
                    os.remove(canonical_path)
                raise HTTPException(status_code=500, detail="Persisted file integrity check failed")

        except Exception as e:
            if os.path.exists(canonical_path):
                os.remove(canonical_path)
            raise HTTPException(status_code=500, detail=f"Failed to verify persistence: {e}")

        try:
            new_template = EbayTemplate(
                original_filename=file.filename,
                file_path=canonical_path,
                file_size=file_size,
                sha256=uploaded_sha256,
                uploaded_at=datetime.utcnow()
            )
            self.db.add(new_template)
            self.db.commit()
            self.db.refresh(new_template)

            print(f"[EBAY_UPLOAD] Success: id={new_template.id} sha256={uploaded_sha256}")
            return new_template

        except Exception as e:
            if os.path.exists(canonical_path):
                os.remove(canonical_path)
            print(f"[EBAY_UPLOAD] DB Error: {e}")
            raise HTTPException(status_code=500, detail="Database error saving template record")

    def _cell_to_string(self, value) -> Optional[str]:
        """
        Deterministic normalization of cell values to string for storage.
        """
        if value is None:
            return None

        if isinstance(value, str):
            s = value.strip()
            return s if s else None

        if isinstance(value, int):
            return str(value)

        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value)

        s = str(value).strip()
        return s if s else None

    def _normalize_field_key(self, s: str) -> str:
        """
        Normalize field name for robust matching:
        - Replace non-breaking spaces
        - Collapse multiple spaces
        - Lowercase
        """
        if not s:
            return ""
        s = s.replace("\u00A0", " ")
        s = re.sub(r"\s+", " ", s)
        return s.strip().lower()

    def _is_header_label(self, s: Optional[str]) -> bool:
        if not s:
            return False
        t = self._normalize_field_key(s)
        return t in ("field names", "field name", "fields", "field")

    def parse_ebay_template(self, template_id: int) -> EbayTemplateParseSummary:
        """
        Parse the stored eBay XLSX template and populate DB fields/values.
        Idempotent: Clears existing fields/values for this template before inserting.
        """
        # 1) Load Template Record
        template = self.db.query(EbayTemplate).filter(EbayTemplate.id == template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")

        if not os.path.exists(template.file_path):
            raise HTTPException(status_code=400, detail=f"Template file missing at {template.file_path}")

        # 2) Load Workbook
        try:
            wb = openpyxl.load_workbook(template.file_path, data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to load Excel file: {e}")

        required_sheets = ["Template", "Valid Values", "Default Values"]
        sheet_names = wb.sheetnames
        for req in required_sheets:
            if req not in sheet_names:
                raise HTTPException(status_code=400, detail=f"Missing required sheet: '{req}'")

        # 3) Clear Existing Data (Idempotency)
        # Delete values first (FK), then fields
        existing_fields = self.db.query(EbayField).filter(EbayField.ebay_template_id == template_id).all()
        existing_field_ids = [f.id for f in existing_fields]

        if existing_field_ids:
            # We don't know the exact column name on the value model across variants, so handle both common cases.
            # Common: ebay_field_id
            if hasattr(self.ValidValueModel, "ebay_field_id"):
                self.db.query(self.ValidValueModel).filter(
                    self.ValidValueModel.ebay_field_id.in_(existing_field_ids)
                ).delete(synchronize_session=False)
            # Alternate: field_id
            elif hasattr(self.ValidValueModel, "field_id"):
                self.db.query(self.ValidValueModel).filter(
                    self.ValidValueModel.field_id.in_(existing_field_ids)
                ).delete(synchronize_session=False)

        self.db.query(EbayField).filter(
            EbayField.ebay_template_id == template_id
        ).delete(synchronize_session=False)

        self.db.flush()

        # 4) Parse Template header row (Template sheet)
        ws_template = wb["Template"]
        header_row_idx = 4  # your convention

        fields_inserted = 0
        field_map_by_key = {}  # normalized_key -> EbayField object

        max_col = ws_template.max_column
        for col_idx in range(1, max_col + 1):
            raw_field_name = self._cell_to_string(ws_template.cell(row=header_row_idx, column=col_idx).value)
            if not raw_field_name:
                continue

            field = EbayField(
                ebay_template_id=template_id,
                field_name=raw_field_name,
                display_name=raw_field_name,
                required=False,
                order_index=col_idx - 1,
                selected_value=None,
                custom_value=None
            )
            self.db.add(field)

            field_map_by_key[self._normalize_field_key(raw_field_name)] = field
            fields_inserted += 1

        # Flush so fields get IDs (needed to insert valid values by FK)
        self.db.flush()

        # 5) Parse Valid Values (supports BOTH formats)
        ws_valid = wb["Valid Values"]
        values_inserted = 0
        values_ignored = 0

        max_row_valid = ws_valid.max_row
        max_col_valid = ws_valid.max_column

        # Determine start row based on whether A1 is a header label
        a1_valid = self._cell_to_string(ws_valid.cell(row=1, column=1).value)
        start_row_valid = 2 if self._is_header_label(a1_valid) else 1

        for r in range(start_row_valid, max_row_valid + 1):
            raw_name = self._cell_to_string(ws_valid.cell(row=r, column=1).value)
            if not raw_name:
                continue
            if self._is_header_label(raw_name):
                continue

            key = self._normalize_field_key(raw_name)
            field = field_map_by_key.get(key)

            if field is None:
                # count actual values we are ignoring on this row
                for c in range(2, max_col_valid + 1):
                    if self._cell_to_string(ws_valid.cell(row=r, column=c).value):
                        values_ignored += 1
                continue

            # Collect values across columns B..N (WIDE format)
            row_values = []
            seen_in_row = set()
            for c in range(2, max_col_valid + 1):
                v = self._cell_to_string(ws_valid.cell(row=r, column=c).value)
                if v and v not in seen_in_row:
                    row_values.append(v)
                    seen_in_row.add(v)

            if not row_values:
                continue

            # Per-field dedupe across all rows
            if not hasattr(field, "_seen_values"):
                field._seen_values = set()

            for v in row_values:
                if v in field._seen_values:
                    continue

                # Insert using whichever FK column exists on your valid-value model.
                if hasattr(self.ValidValueModel, "ebay_field_id"):
                    self.db.add(self.ValidValueModel(ebay_field_id=field.id, value=v))
                elif hasattr(self.ValidValueModel, "field_id"):
                    self.db.add(self.ValidValueModel(field_id=field.id, value=v))
                else:
                    # As a last resort, try relationship append if it exists
                    if hasattr(field, "valid_values"):
                        field.valid_values.append(self.ValidValueModel(value=v))
                    else:
                        raise RuntimeError("Valid value model has no FK field_id/ebay_field_id and field has no valid_values relationship.")

                field._seen_values.add(v)
                values_inserted += 1

        # 6) Parse Default Values
        ws_defaults = wb["Default Values"]
        defaults_applied = 0
        defaults_ignored = 0

        max_row_def = ws_defaults.max_row
        a1_def = self._cell_to_string(ws_defaults.cell(row=1, column=1).value)
        start_row_def = 2 if self._is_header_label(a1_def) else 1

        for r in range(start_row_def, max_row_def + 1):
            raw_name = self._cell_to_string(ws_defaults.cell(row=r, column=1).value)
            if not raw_name or self._is_header_label(raw_name):
                continue

            default_val = self._cell_to_string(ws_defaults.cell(row=r, column=3).value)
            if not default_val:
                continue

            key = self._normalize_field_key(raw_name)
            field = field_map_by_key.get(key)

            if field is not None:
                field.selected_value = default_val
                defaults_applied += 1
            else:
                defaults_ignored += 1

        # Commit
        self.db.commit()

        return EbayTemplateParseSummary(
            template_id=template_id,
            fields_inserted=fields_inserted,
            values_inserted=values_inserted,
            defaults_applied=defaults_applied,
            values_ignored_not_in_template=values_ignored,
            defaults_ignored_not_in_template=defaults_ignored,
            sheet_names=sheet_names
        )
